#  Working envrionment: Python 2.7
from openpyxl import load_workbook
from openpyxl.styles import colors,PatternFill
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdftypes import resolve1
from pdfminer.psparser import PSLiteral, literal_name
style = PatternFill(fill_type='solid',fgColor=colors.RED)
import os

#  extract data from pdf-acroform fields
def load_fields_from_pdf(field, T=''):
    #  Recursively load form fields
    form = field.get('Kids', None)
    t = field.get('T')
    if t is None:
        t = T
    else:
        #  Add its father name
        t = T + '.' + t if T != '' else t
    """ Following is to repeat fields that have "Kids", now is commented because 
    1. There could be multiple fileds who shared the same field name.
    2. For buttons, the parents has "V" value already, don't need to dig in Kids.
    """
    # if form and t:
    #     return [load_fields_from_pdf(resolve1(f), t) for f in form]
    # else:
    # Some field types, like signatures, need extra resolving
    value = resolve1(field.get('AS')) if resolve1(field.get('AS')) is not None else resolve1(field.get('V'))
    #  if output is PSLiteral type, transfer it into str type through "literal_name" function
    if isinstance(value, PSLiteral):
        return (t, literal_name(value))
    else:
        return (t, resolve1(value))


#  split data into dictionary
def split_data(field, d={}):
    flag = True if len(field) != 2 else False
    if flag:
        for f in field:
            split_data(f, d)
    elif isinstance(field[0], (tuple, list)):
        for f in field:
            split_data(f, d)
    else:
        key = field[0] if field[0] is not None else field[0]
        d[key] = field[1]
    return d

#  load ICS data from decrypted pdf docutment
def load_data_from_pdf(pdf):
    with open(pdf, 'rb') as file:
        parser = PDFParser(file)
        doc = PDFDocument(parser)
        parser.set_document(doc)
        outcome = [load_fields_from_pdf(resolve1(f)) for f in resolve1(doc.catalog['AcroForm'])['Fields']]
    # format the outcome of data extract from ics pdf
    outcome = split_data(outcome)
    if outcome['Max Dynamic Reader Limit sets supported']:
        outcome['Max Dynamic Reader Limit sets supported'] = 'Yes' if int(outcome['Max Dynamic Reader Limit sets supported']) > 4 else 'No'
    if outcome['Product Configuration']:
        outcome['Product Configuration'] = 'Yes'  if outcome['Product Configuration'] == '(A) PCDA (IRWIN Reader) / S-ICR' else 'No'
    for key in outcome:
        if outcome[key] == 'Off':
            outcome[key] = 'No'
    return outcome

#  load ICS data from xlsx document
def load_data_from_xlsx(xlsx):
    dic = {}
    try:
        ics = load_workbook(xlsx)
    except IOError as e:
        raw_input(e)
        exit()
    ics_sheet = ics.active
    for cell in ics_sheet['1']:
        ics_key = cell.value
        dic[ics_key] = ics_sheet[cell.column + '2'].value
    ics.close()
    return dic

#  fill ICS data into the latest VISA template
def fill_in_data(ics_dic, tem):
    #  fill in the data information
    try:
        template = load_workbook(tem)
    except IOError as e:
        raw_input(e)
        exit()
    sheet_template = template.get_sheet_by_name('ICS')
    for formCell in sheet_template['E']:
        formID = formCell.value
        if formID is not None:
            if formID in ics_dic.keys():
                output = ics_dic[formID]
                #  the output type of pdf is unicode/string, need to be took into consider
                if formID == 'Visa Contactless Reader Implementation Notes Version' and output not in [None, '']:
                    output = 'Yes' if output in [1.1, '1.1'] else 'No'
                if formID == 'Max Dynamic Reader Limit sets supported':
                    if ics_dic['Check Box30'] == 'No':
                        output = 'No'
                # # due to ics pdf template is wrongly implemented, we have temporary solution as below to define yes for 3 specific question
                # if formID in ['JlX2RldmljZV9WQ1BTMi4yLmh0bWwA.section7.zeroamtauth_(1)',
                #               'JlX2RldmljZV9WQ1BTMi4yLmh0bWwA.section7.approved_receipt',
                #               'JlX2RldmljZV9WQ1BTMi4yLmh0bWwA.section7.declined_receipt'] and output in ['', None]:
                #     output = 'Yes'
                #  if there is any information other than Yes or No, mark the cell as RED
                if output not in ['Yes', 'No']:
                    sheet_template['G' + str(formCell.row)].fill = style
                else:
                    sheet_template['G' + str(formCell.row)].value = output
    #  check if there is any blank cells to fill, if yes, mark as RED
    for ref in sheet_template['A']:
        if ref.value not in [None, 'Reference']:
            if sheet_template['G' + str(ref.row)].value is None:
                sheet_template['G' + str(ref.row)].fill = style
            #  check for template 2.2
            if tem == 'Template_22.xlsx' and sheet_template['C' + str(ref.row)].value in [2.3, 5.1, 5.2, 5.3, 5.7]:
                if sheet_template['G' + str(ref.row)].value != 'No':
                    sheet_template['G' + str(ref.row)].fill = style
            if tem == 'Template_22.xlsx' and sheet_template['C' + str(ref.row)].value == 3.14:
                if sheet_template['G' + str(ref.row)].value == 'No' and sheet_template['G' + str(ref.row + 1)].value == 'Yes':
                    sheet_template['G' + str(ref.row)].fill = style
    template.save(os.path.dirname(__file__) + 'expectResult.xlsx')
    template.close()
    print 'Fill in data succeeded!\n'

#  compare output template with manual one
def compare(exl_b):
    dic_a = {}
    dic_b = {}
    diff = []
    try:
        file_a = load_workbook('expectResult.xlsx')
        file_b = load_workbook(exl_b)
    except IOError as e:
        raw_input(e)
        exit()
    sheet_a = file_a.get_sheet_by_name('ICS')
    sheet_b = file_b.get_sheet_by_name('ICS')
    id_a = [find_title.column for find_title in sheet_a['1'] if find_title.value == 'Question ID'][0]
    value_a = [find_title.column for find_title in sheet_a['1'] if find_title.value == 'Value'][0]
    id_b = [find_title.column for find_title in sheet_b['1'] if find_title.value == 'Question ID'][0]
    value_b = [find_title.column for find_title in sheet_b['1'] if find_title.value == 'Value'][0]
    for cell in sheet_a[id_a]:
        key_a = cell.value
        dic_a[key_a] = sheet_a[value_a + str(cell.row)].value
    for cell in sheet_b[id_b]:
        key_b = cell.value
        dic_b[key_b] = sheet_b[value_b + str(cell.row)].value
    #  compare two dictionary
    if cmp(dic_a, dic_b) == 0:
        print 'Results are the same.'
    elif cmp(dic_a.keys(), dic_b.keys()) != 0:
        print 'Warning! There are different ICS Questions between two templates, please check ICS Template version: \n', set(dic_a) ^ set(dic_b)
    else:
        for i in dic_a.keys():
            for j in dic_b.keys():
                if i == j:
                    if dic_a[i] == dic_b[j]:
                        pass
                    else:
                        diff.append(i)
        print 'The following items(Question ID) are not the same, please check and modify correctly\n', diff
    file_a.close()
    file_b.close()


def main():
    action = raw_input('------ What service you want to use with checktool ------\n'
                       '1. Generate template from ICS\n2. Compare difference with two template excel\n3. Exit\n')
    if action == '1':
        ics_dic = {}
        temp = 'Template_213.xlsx' if raw_input('------ Please select RRAP Temaplate Version ------\n'
                                                '1. 2.1.3\n2. 2.2\n') == '1' else 'Template_22.xlsx'
        source = raw_input('------ What source do you want to input ICS ------\n1. pdf\n2. xlsx\n')
        if source == '1':
            path = raw_input('------ Please input the path of the pdf file. ------\n'
                             'Note: The pdf file must be decrypted.\n').strip()
            ics_dic = load_data_from_pdf(path)
        if source == '2':
            path = raw_input('------ Please input the path of the xlsx file. ------\n'
                             'Note: You can generate xlsx file through Adobe Acrobat 7.0.\n').strip()
            ics_dic = load_data_from_xlsx(path)
        fill_in_data(ics_dic, temp)
    if action == '2':
        f = raw_input('------ Please input the path of manual RRAP Excel. ------\n'
                           'Note: Please remove "Data Validation Format" before comparing the files!\n').strip()
        compare(f)
    if action == '3':
        exit()
    return main()


#  Only for debug use
def test():
    pass
    #ics_dic1 = load_data_from_pdf('out1.pdf')
    # a = {123:123, 23:23, }
    # b = {123:123, 23:24,}
    # print set(b) - set(a)
    # print set(a) - set(b)
    # ics_dic1 = load_data_from_pdf('out1.pdf')
    # ics_dic2 = load_data_from_xlsx('reportmag.xlsx')
    # print ics_dic1
    # print ics_dic2
    #fill_in_data(ics_dic1, 'template_22.xlsx', 'out1.pdf')
    #compare(1,2)
    # if cmp(ics_dic1, ics_dic2) == 0:
    #     print 'Results are the same.'
    # if cmp(ics_dic1.keys(),ics_dic2.keys()) != 0:
    #     print 'Warning! There are different ICS questions between two template, please update as following item:', set(ics_dic2) ^ set(ics_dic1)

if __name__ == '__main__':
    #test()
    main()